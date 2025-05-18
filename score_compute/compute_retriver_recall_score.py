import json
import argparse
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# 读取JSON文件
def calculate_ap_at_k(pred, gt, k=10):
    
    k = min(k, len(pred))

    hits = 0
    avg_precision = 0.0
    for rank, item in enumerate(pred[:k], start=1):  # 遍历前 k 个预测结果
        if item in gt:  # 如果预测结果在 GT 中
            hits += 1
            avg_precision += hits / rank  # 累加 Precision@i

    if hits == 0:
        return 0.0  # 如果没有匹配到相关项，AP 为 0
    return avg_precision / len(gt)  # 平均精度归一化到 GT 中的相关项数量 


def F1_score(label_list, result_list):
    TP = 0
    for pred in result_list:
        if pred in label_list:
            TP += 1
    if TP == 0:
        return 0,0,0
    precision = TP / len(result_list)
    # print(precision)
    recall = TP / len(label_list)
    F1 = 2 * precision * recall / (precision+recall)
    if F1>1:
        F1=1
    if precision>1:
        precision=1
    if recall>1:
        recall=1
    return F1, precision, recall


def compute_recall_map_and_save_csv(top_k,data):
    

    recall_score_all=0
    count=0
   


    recall_score_all_sp=0
    count_sp=0

    recall_score_all_mp=0
    count_mp=0


    recall_layout = 0
    recall_text = 0
    
    recall_vis = 0

    count_text = 0
    count_layout = 0
    count_illu = 0
    count_other = 0
    count_vis = 0

    recall_score_all_mh = 0
    count_mh = 0

    for key, value in data.items():
        for item in value:
            evidence_pages = item.get('evidence_pages', None)
            top_page = item.get('top_pages', None)
            top_page = [int(num) for num in top_page]
            if evidence_pages:
                
                evidence_pages=json.loads(evidence_pages)
                
            #     ]
            if evidence_pages and top_page:
                
                F1, precision, recall=F1_score(evidence_pages,top_page[0:top_k])
                
                ap_score = calculate_ap_at_k(top_page, evidence_pages, k=top_k)
                

                item['F1_score'] = F1
                item['precision'] = precision
                item['recall'] = recall
                item["ap_score"] = ap_score

                

                if item["type"]=="单页问答":
                   
                    recall_score_all_sp += recall
                   
            
                    count_sp += 1
                    
                elif item["type"]=="跨页问答":
                   
                    recall_score_all_mp += recall
                   
                    count_mp += 1

                elif item["type"]=="多跳问答":
                    recall_score_all_mh += recall

                    count_mh += 1

                if item["evidence_modal"]=="text":
                    
                    
                    if item["type"]=="单页问答" or item["type"]=="跨页问答" or item["type"]=="多跳问答":  
                        recall_text += recall
                            
                
                        count_text += 1

                if item["evidence_modal"]=="layout":

                    if item["type"]=="单页问答" or item["type"]=="跨页问答" or item["type"]=="多跳问答":  
                        recall_layout += recall
                        
            
                        count_layout += 1

                if item["evidence_modal"]=="illustration" or item["evidence_modal"]=="其他":

                    if item["type"]=="单页问答" or item["type"]=="跨页问答" or item["type"]=="多跳问答":  
                        recall_vis += recall
                        
            
                        count_vis += 1
                
               
                if item["type"]=="单页问答" or item["type"]=="跨页问答" or item["type"]=="多跳问答":
                    
                    recall_score_all = recall_score_all + recall
                   
                    count=count+1

   
    recall_score_average=recall_score_all/count*100

    
    recall_score_average_sp = recall_score_all_sp/count_sp*100
    
    recall_score_average_mp = recall_score_all_mp/count_mp*100
  
    recall_score_average_mh = recall_score_all_mh/count_mh*100

    

    avg_recall_layout = recall_layout/count_layout*100
    
    avg_recall_vis = recall_vis/count_vis*100
    avg_recall_text = recall_text/count_text*100

    

    return recall_score_average_sp,recall_score_average_mp, recall_score_average_mh,avg_recall_layout, avg_recall_vis,avg_recall_text,recall_score_average


if __name__ == "__main__":

    input_json_file='/data1/szy/LLM_code/GujiDoc_collections/LCHD_benchmark_test_code/retriver_res/bge-m3_retrival_res.json' #Input the path to the retriver inference result JSON for scoring.

    with open(input_json_file, 'r') as f:
        data = json.load(f)

    file_path, file_ext = os.path.splitext(input_json_file)
    output_json_file = f"{file_path}_score{file_ext}"

    

    top_k_ls = [1,2,5,25] 
    data_dic={}
    print(f"result:{output_json_file}")
    for top_k in top_k_ls:
        recall_score_average_sp,recall_score_average_mp, recall_score_average_mh,avg_recall_layout, avg_recall_vis,avg_recall_text,recall_score_average = compute_recall_map_and_save_csv(top_k,data)
        

        data_dic[f"recall_average_sp_{top_k}"]=round(recall_score_average_sp,2)
        data_dic[f"recall_average_mp_{top_k}"]=round(recall_score_average_mp,2)
        data_dic[f'recall_average_mh_{top_k}']=round(recall_score_average_mh,2)

        data_dic[f"recall_average_text_{top_k}"]=round(avg_recall_text,2)
        data_dic[f"recall_average_layout_{top_k}"]=round(avg_recall_layout,2)
       
        data_dic[f"recall_average_vis_{top_k}"]=round(avg_recall_vis,2)
    
    

        data_dic[f"recall_score_average_{top_k}"]=round(recall_score_average,2)



    

    # 生成数据
    model_name = input_json_file.split("/")[-1].replace(".json","")
    data = [
        ['','单页','单页', '双页', '双页', '多跳','多跳','text','layout','visual','综合'],
        ['top_k','top_1', 'top_5','top_2','top_5','top_5','top_25','top_5','top_5','top_5','top_5'],
        [model_name,data_dic['recall_average_sp_1'], data_dic['recall_average_sp_5'], data_dic['recall_average_mp_2'], data_dic['recall_average_mp_5'], data_dic['recall_average_mh_5'], data_dic['recall_average_mh_25'],data_dic['recall_average_text_5'],data_dic['recall_average_layout_5'],data_dic['recall_average_vis_5'],data_dic['recall_score_average_5']]
    ]


   
    df = pd.DataFrame(data)

 
    wb = Workbook()
    ws = wb.active

    ws.title = model_name


    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)


    column_widths = [10] * len(df.columns)  
    for col, width in zip(ws.columns, column_widths):
        max_length = 0
        column = col[0].column_letter  
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 合并单元格
    ws.merge_cells('B1:C1')  
    ws.merge_cells('D1:E1')  
    ws.merge_cells('F1:G1')  
    ws.merge_cells('H2:K2')  
    

    # 设置单元格内容对齐方式
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=14):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 保存 Excel 文件
    wb.save(f"./retriver_res/xlsx/{model_name}.xlsx")
    print(f"./retriver_res/xlsx/{model_name}.xlsx")




