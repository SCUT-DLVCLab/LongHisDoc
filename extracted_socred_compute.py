import json
import argparse
import os
import re
from zhconv import convert
import ast
# import Levenshtein

# 读取JSON文件


def levenshtein_distance(s1, s2):
    if len(s1) > len(s2):
        s1, s2 = s2, s1

    distances = range(len(s1) + 1)
    for i2, c2 in enumerate(s2):
        distances_ = [i2 + 1]
        for i1, c1 in enumerate(s1):
            if c1 == c2:
                distances_.append(distances[i1])
            else:
                distances_.append(1 + min((distances[i1], distances[i1 + 1], distances_[-1])))
        distances = distances_
    return distances[-1]

def custom_trad_to_simp(text):
    mapping = {"𡊮": "袁","叅":"参"}
    for trad, simp in mapping.items():
        text = text.replace(trad, simp)
    text = convert(text, 'zh-cn')
    return text

def levenshtein_distance(s1, s2):
    if len(s1) > len(s2):
        s1, s2 = s2, s1

    distances = range(len(s1) + 1)
    for i2, c2 in enumerate(s2):
        distances_ = [i2 + 1]
        for i1, c1 in enumerate(s1):
            if c1 == c2:
                distances_.append(distances[i1])
            else:
                distances_.append(1 + min((distances[i1], distances[i1 + 1], distances_[-1])))
        distances = distances_
    return distances[-1]


def anls_compute(prediction, groundtruth,threshold=0.5):
    if isinstance(prediction,list):
        prediction = prediction[0]
    prediction = custom_trad_to_simp(prediction)
    groundtruth = custom_trad_to_simp(groundtruth)
    # prediction = convert(prediction, 'zh-cn')
    # groundtruth = convert(groundtruth, 'zh-cn')
     
    dist = levenshtein_distance(groundtruth, prediction)
    length = max(len(groundtruth.upper()), len(prediction.upper()))
    value = 0.0 if length == 0 else float(dist) / float(length)
    anls = 1.0 - value
    if anls<=threshold:
        anls = 0.0
    return anls

def score_int(pred, ref):
    # pred = custom_trad_to_simp(pred)
    # ref = custom_trad_to_simp(ref)
    return 1.0 if pred == ref else 0.0


# def get_clean_string(s):
#     s = str(s).lower().strip()
#     s = s.replace("，", "")
#     s = s.replace("。", "")
#     s = s.replace("。","")
#     # remove parenthesis
#     # s = re.sub(r'\s*\([^)]*\)', '', s).strip()
#     # # remove quotes
#     # s = re.sub(r"^['\"]|['\"]$", "", s).strip()
#     # s = s.strip().lstrip("$").strip()
#     # s = s.strip().lstrip("£").strip()
#     # s = s.strip().rstrip("%").strip()
#     return s




def compute_mean(score_ls):
    mean = sum(score_ls)/len(score_ls)*100
    return mean



input_json_file='/data1/szy/LLM_code/GujiDoc_collections/LCHD_benchmark_test_code/evidence_input/Qwen2.5-VL-7B-Instruct_evidence_5_10.json'

output_file_path = input_json_file


with open(input_json_file, 'r') as f:
    guji_data = json.load(f)

score_all_ls = []
for guji_id in guji_data:
    guji_book = guji_data[guji_id]
    
    for data in guji_book:
        gt = data["answer"]
        # import pdb
        # pdb.set_trace()
        # pred_ans = data["model_res"]
        if "model_res" not in data:
            # import pdb
            # pdb.set_trace()
            data["score"] = 0.0
        elif "LLM_scored" not in data:
            data["score"] = 0.0
        # if "model_res" not in data:
        #     data["score"] = 0.0
            # if "score" in data:
            #     score_all_ls.append(data["score"])
            # continue
            # import pdb
            # pdb.set_trace()
        else:
            if data["model_res"]==None:
                data["score"] = 0.0
            else:
                if data["answer_type"]=="Str":
                    model_res = data["model_res"]
                    model_res = model_res.rstrip()
                    score1 = anls_compute(model_res,gt)
                   
                    # import pdb
                    # pdb.set_trace()
                    score2 = anls_compute(data["LLM_scored"], gt)
                    LLM_scored = data["LLM_scored"]

                    if score2 >= score1:
                        data["score"] = score2
                    else:
                        print(f"gt：{gt}")
                        print(f"model_res:{model_res}")
                        print(f"LLM_scored:{LLM_scored}")
                        print()
                        data["score"] = score1


                elif data["answer_type"]=="Bool" :
                    # LLM_score = data["LLM_scored"]
                    # except:
                    #     import pdb
                    #     pdb.set_trace()

                    if data["LLM_scored"]==gt or data["model_res"]==gt:
                        data["score"] = 1
                    else:
                        data["score"] = 0

                elif data["answer_type"]=="Enum":
                    if data["LLM_scored"].upper()==gt.upper() or data["model_res"].upper()==gt.upper() :
                        data["score"] = 1
                    else:
                        data["score"] = 0


                elif data["answer_type"]=="None":
                    if "无法回答" in data["LLM_scored"]:
                        data["score"] = 1
                    else:
                        data["score"] = 0
                    

                elif data["answer_type"]=="Int":
                    try:
                        pred_int = data["LLM_scored"]
                        if isinstance(gt,list):
                            if str(pred_int) in gt:
                                data["score"] = 1.0

                            else:
                                # import pdb
                                # pdb.set_trace()
                                data["score"] = 0.0
                        else:
                            if pred_int == int(gt):
                                data["score"] = 1.0
                            else:
                                # pdb.set_trace()
                                data["score"] = 0.0
                    except:
                        # import pdb
                        # pdb.set_trace()
                        data["score"] = 0.0

                


                elif data["answer_type"]=="List":
                    # import pdb
                    # pdb.set_trace()
                    if isinstance(data["LLM_scored"],list):
                        pred_ls = data["LLM_scored"]
                        
                    
                    # elif:
                    else:
                        try:
                            # import pdb
                            # pdb.set_trace()
                            # pred_ls = json.loads(data["LLM_scored"])
                            pred_ls = ast.literal_eval(data["LLM_scored"])
                
                        except:
                            pred_ls =[]
                            print("无法提取")
                            # import pdb
                            # pdb.set_trace()
                            data["score"] = 0.0
                            score_all_ls.append(data["score"])
                            continue
                            
                            # import pdb
                            # pdb.set_trace()

                    try:
                        if all(isinstance(item, int) for item in pred_ls):
                            if not pred_ls or not gt:
                                final_score = 0.0  # 避免除零错误
                            
                            else:
                                greedy_scores_list = [
                                    max(score_int(pred, ref) for pred in pred_ls) for ref in gt
                                ]
                                
                                final_score = sum(greedy_scores_list) / len(gt)* min(1, len(gt) / len(pred_ls)) ** 0.5
                        else:
                            if not pred_ls or not gt:
                                final_score = 0.0  # 避免除零错误
                            
                            else:
                                # if gt[0] == "茶之源":
                                    # import pdb
                                    # pdb.set_trace()
                                greedy_scores_list = [
                                    max(anls_compute(pred, ref) for pred in pred_ls) for ref in gt
                                ]
                                

                                final_score = sum(greedy_scores_list) / len(gt)* min(1, len(gt) / len(pred_ls)) ** 0.5

                        data["score"] = final_score
                    except:
                        data["score"] = 0.0
                    

        if "score" in data:
            score_all_ls.append(data["score"])
import pdb
pdb.set_trace()
mean = sum(score_all_ls)/len(score_all_ls)*100
# import pdb
# pdb.set_trace()

##分类别统计分数
sp_ls = []
cp_ls = []
mh_ls = []
ua_ls = []


text_ls = []
layout_ls = []
illu_ls = []
other_ls = []
vis_ls = []


str_ls = []
bool_ls = []
enum_ls = []
int_ls = []
list_ls = []
none_ls = []


for guji_id in guji_data:
    guji_book = guji_data[guji_id]
    
    for data in guji_book:
        # if "LLM_scored" not in data:
        #     continue

        if data["type"] == "单页问答":
            sp_ls.append(data["score"])

        elif data["type"] == "跨页问答":
            cp_ls.append(data["score"])
            

        elif data["type"] == "多跳问答":
            mh_ls.append(data["score"])

        elif data["type"] == "无法回答":
            ua_ls.append(data["score"])


        if data["evidence_modal"] == "text":
            text_ls.append(data["score"])

        elif data["evidence_modal"] == "layout":
            layout_ls.append(data["score"])

        elif data["evidence_modal"] == "illustration" or data["evidence_modal"] == "其他":
            vis_ls.append(data["score"])

        # elif data["evidence_modal"] == "illustration":
        #     illu_ls.append(data["score"])
        
        # elif data["evidence_modal"] == "其他":
        #     other_ls.append(data["score"])


        if data["answer_type"]=="Str":
            str_ls.append(data["score"])
        
        elif data["answer_type"]=="Bool":
            bool_ls.append(data["score"])

        elif data["answer_type"]=="Enum":
            enum_ls.append(data["score"])

        elif data["answer_type"]=="Int":
            int_ls.append(data["score"])

        elif data["answer_type"]=="List":
            list_ls.append(data["score"])

        elif data["answer_type"]=="None":
            none_ls.append(data["score"])

sp_mean = round(compute_mean(sp_ls),2)

cp_mean = round(compute_mean(cp_ls),2)

mh_mean = round(compute_mean(mh_ls),2)

# ua_mean = ""

ua_mean = compute_mean(ua_ls)

text_mean = round(compute_mean(text_ls),2)

layout_mean = round(compute_mean(layout_ls),2)

# illu_mean = round(compute_mean(illu_ls),2)

# other_mean = round(compute_mean(other_ls),2)

vis_ls_mean = round(compute_mean(vis_ls),2)

str_mean = round(compute_mean(str_ls),2)

bool_mean = round(compute_mean(bool_ls),2)

enum_mean = round(compute_mean(enum_ls),2)

int_mean = round(compute_mean(int_ls),2)

list_mean = round(compute_mean(list_ls),2)

none_mean = compute_mean(none_ls)


import pdb
pdb.set_trace()


with open(output_file_path , 'w', encoding='utf-8') as json_file:
    json.dump(guji_data, json_file, ensure_ascii=False, indent=4)



import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# 生成数据
setting_name = input_json_file.split("/")[-1].replace(".json","")
data = [
    ['','单页','跨页','多跳','无法回答','Str','Int','Enum','LIST','Bool','None','text','layout','vis','None','综合'],
    [setting_name, sp_mean , cp_mean , mh_mean , ua_mean,str_mean,int_mean,enum_mean,list_mean,bool_mean,none_mean,text_mean,layout_mean, vis_ls_mean,none_mean,mean]
]
# data = [
#     ['','单页','跨页','多跳','无法回答','Str','Int','Enum','LIST','Bool','None','text','layout','illu', 'other','None','综合'],
#     [setting_name, sp_mean , cp_mean , mh_mean , ua_mean,str_mean,int_mean,enum_mean,list_mean,bool_mean,none_mean,text_mean,layout_mean, illu_mean, other_mean,none_mean,mean]
# ]
# 将数据转化为 DataFrame
df = pd.DataFrame(data)

# 创建一个新的 Excel 工作簿
wb = Workbook()
ws = wb.active

ws.title = setting_name

# 将 DataFrame 写入 Excel
for row in dataframe_to_rows(df, index=False, header=False):
    ws.append(row)

# 设置列宽
column_widths = [10] * len(df.columns)  # 设置每列的宽度
for col, width in zip(ws.columns, column_widths):
    max_length = 0
    column = col[0].column_letter  # 获取列名
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# 合并单元格
# ws.merge_cells('B1:C1')  # 第一行单页占8个单元格
# ws.merge_cells('D1:E1')  # 第一行双页占6个单元格
# ws.merge_cells('F2:H2')  # 第二行top_1
# ws.merge_cells('D2:E2') 
# ws.merge_cells('F2:G2') 
# ws.merge_cells('H2:I2') 
# ws.merge_cells('J2:K2') 
# ws.merge_cells('L2:M2')  # 第二行top_2, top_5, top_10
# ws.merge_cells('N2:O2')  # 第二行top_2, top_5, top_10

# 设置单元格内容对齐方式
for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=14):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存 Excel 文件
xlsx_file = f"/data1/szy/LLM_code/GujiDoc_collections/LCHD_benchmark_test_code/res/xlsx/{setting_name}.xlsx"
if os.path.exists(xlsx_file):
    print("已存在xlsx")
wb.save(xlsx_file)
print(xlsx_file)




